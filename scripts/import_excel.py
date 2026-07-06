#!/usr/bin/env python3
"""
Import weeks 1-4 workout data from Roberto_Fitness_Tracker_v2.xlsx into Supabase.

Requires environment variables:
  SUPABASE_URL          - e.g. https://yrihprcinsramxebvilf.supabase.co
  SUPABASE_SERVICE_KEY  - service role key (NEVER commit or expose in frontend)

Usage:
  set SUPABASE_URL=https://yrihprcinsramxebvilf.supabase.co
  set SUPABASE_SERVICE_KEY=your-service-role-key
  python scripts/import_excel.py --user-id <uuid> --excel "C:\\Users\\User\\Downloads\\Roberto_Fitness_Tracker_v2.xlsx"

The script creates workout_sessions and workout_sets for logged weeks 1-4.
Dates are inferred from the program start date (2026-06-01) and weekly schedule.
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import date, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    from supabase import create_client
except ImportError:
    print("Install supabase: pip install supabase", file=sys.stderr)
    sys.exit(1)

WORKOUT_SHEETS = ["Push A", "Pull A", "Legs A", "Push B", "Pull B", "Legs B"]
DEFAULT_SCHEDULE = ["Rest", "Push A", "Pull A", "Legs A", "Push B", "Pull B", "Legs B"]
START_DATE = date(2026, 6, 1)
MAX_WEEK = 4


def parse_num(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", ".")
    m = re.match(r"^[\d.]+", s)
    return float(m.group()) if m else None


def parse_int(val):
    n = parse_num(val)
    return int(n) if n is not None else None


def dates_for_workout(week: int, routine_name: str, start: date = START_DATE) -> date:
    """Session date for routine_name in the given program week (1-indexed)."""
    week_start = start + timedelta(days=(week - 1) * 7)
    for offset in range(7):
        d = week_start + timedelta(days=offset)
        schedule_index = (d.weekday() + 1) % 7  # Mon=1 .. Sun=0
        if DEFAULT_SCHEDULE[schedule_index] == routine_name:
            return d
    return week_start


def extract_sets(row):
    """Parse Set 1 W/R through Set 4 W/R from workout log row."""
    sets = []
    # columns 6-13: Set1W, Set1R, Set2W, Set2R, Set3W, Set3R, Set4W, Set4R
    for i in range(4):
        w = parse_num(row[6 + i * 2])
        r = parse_int(row[7 + i * 2])
        if w is not None and r is not None and r < 100:  # skip bad cells like 94.0 typo reps
            sets.append({"set_number": i + 1, "weight_kg": w, "reps": r})
    rir = parse_num(row[14]) if len(row) > 14 else None
    if rir is not None and sets:
        sets[-1]["rir"] = rir
    return sets


def load_sheet(wb, sheet_name: str, max_week: int = MAX_WEEK):
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=5, values_only=True))
    entries = []
    current_week = None
    for row in rows:
        if not row or not row[0]:
            continue
        week_val = parse_num(row[0])
        if week_val is not None:
            current_week = int(week_val)
        if current_week is None or current_week > max_week:
            continue
        exercise = row[2]
        if not exercise:
            continue
        sets = extract_sets(row)
        if not sets:
            continue
        entries.append({
            "week": current_week,
            "exercise": str(exercise).strip(),
            "sets": sets,
            "target_sets": row[3],
            "notes": row[15] if len(row) > 15 else None,
        })
    return entries


def main():
    parser = argparse.ArgumentParser(description="Import Excel workout logs into Supabase")
    parser.add_argument("--user-id", required=True, help="Supabase auth user UUID")
    parser.add_argument("--excel", default=r"C:\Users\User\Downloads\Roberto_Fitness_Tracker_v2.xlsx")
    parser.add_argument("--weeks", type=int, default=MAX_WEEK, help="Max week to import (default 4)")
    args = parser.parse_args()

    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_KEY")
    if not url or not key:
        print("Set SUPABASE_URL and SUPABASE_SERVICE_KEY environment variables.", file=sys.stderr)
        sys.exit(1)

    if not Path(args.excel).exists():
        print(f"Excel file not found: {args.excel}", file=sys.stderr)
        sys.exit(1)

    sb = create_client(url, key)
    wb = openpyxl.load_workbook(args.excel, data_only=True)

    # Load routine templates + exercises for user
    routines = sb.table("routine_templates").select("id, name").eq("user_id", args.user_id).execute().data or []
    routine_map = {r["name"]: r["id"] for r in routines}
    exercises_by_routine = {}
    for r in routines:
        exs = sb.table("routine_exercises").select("id, name").eq("routine_id", r["id"]).execute().data or []
        exercises_by_routine[r["name"]] = {e["name"]: e["id"] for e in exs}

    imported_sessions = 0
    imported_sets = 0

    for sheet in WORKOUT_SHEETS:
        if sheet not in wb.sheetnames:
            print(f"Skip missing sheet: {sheet}")
            continue
        entries = load_sheet(wb, sheet, args.weeks)
        by_week = {}
        for e in entries:
            by_week.setdefault(e["week"], []).append(e)

        for week, week_entries in by_week.items():
            session_date = dates_for_workout(week, sheet).isoformat()
            routine_id = routine_map.get(sheet)

            session_row = {
                "user_id": args.user_id,
                "routine_id": routine_id,
                "routine_name": sheet,
                "session_date": session_date,
                "week_number": week,
                "completed_at": f"{session_date}T18:00:00Z",
            }
            res = sb.table("workout_sessions").upsert(
                session_row,
                on_conflict="user_id,session_date,routine_name",
            ).execute()
            session = res.data[0] if res.data else None
            if not session:
                existing = sb.table("workout_sessions").select("id").eq("user_id", args.user_id).eq("session_date", session_date).eq("routine_name", sheet).execute()
                session = existing.data[0] if existing.data else None
            if not session:
                print(f"Failed to create session: {sheet} week {week}")
                continue
            session_id = session["id"]
            imported_sessions += 1

            ex_map = exercises_by_routine.get(sheet, {})
            for entry in week_entries:
                for s in entry["sets"]:
                    set_row = {
                        "session_id": session_id,
                        "exercise_id": ex_map.get(entry["exercise"]),
                        "exercise_name": entry["exercise"],
                        "set_number": s["set_number"],
                        "weight_kg": s["weight_kg"],
                        "reps": s["reps"],
                        "rir": s.get("rir"),
                    }
                    sb.table("workout_sets").insert(set_row).execute()
                    imported_sets += 1

            print(f"Imported {sheet} week {week} ({session_date}): {len(week_entries)} exercises")

    print(f"Done. Sessions: {imported_sessions}, sets: {imported_sets}")


if __name__ == "__main__":
    main()
